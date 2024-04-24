from selenium import webdriver
import time
import csv
import requests
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller

import requests
import re
import traceback
from bs4 import BeautifulSoup

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'
}

from openpyxl import load_workbook, Workbook

sheet_name = 'data.xlsx'

# Write Headline and create a new excel sheet
def xl_sheet_headlines(sheet_name=sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = ['url', 'name', 'address', 'website', 'phone',  'email']
    ws.append(headlines)
    wb.save(sheet_name)

xl_sheet_headlines()

# Write Data On existing sheet
def xl_write(data_write, sheet_name=sheet_name):
    wb = load_workbook(sheet_name)
    work_sheet = wb.active  # Get active sheet
    work_sheet.append(data_write)
    wb.save(sheet_name)

def driver_define():
    print('Chromedriver Installing')
    driver_path = chromedriver_autoinstaller.install()

    print('Chrome Browser Opening')
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    s = Service(driver_path)
    driver = webdriver.Chrome(service=s, options=options)
    return driver

# Email Get
def get_email(url):
    domain = url.split('//')[-1].replace('www.', '').split('/')[0]
    url_gen = f'http://www.skymem.info/srch?q={domain}'
    response = requests.get(url_gen, headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    email_list = re.findall(r"href=\"\/srch\?q=(.*?@.*)\">", str(soup))
    email = [line for line in email_list if domain in line]
    if email:
        email = email[0]
    else:
        email = ''

    return email

def scrape_data(driver, url):
    driver.get(url)

    try:
        name_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//div[@class="TIHn2 "]//h1[@class="DUwDvf lfPIob"]')))
        name = name_element.text
    except:
        name = ''

    try:
        address = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[@data-item-id="address"]'))).text
    except:
        address = ''

    try:
        website = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[aria-label^="Website:"]'))).get_attribute('href')
    except:
        website = ''

    try:
        phone = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[aria-label*="Phone:"]'))).text
    except:
        phone = ''

    email = ''
    try:
        if website != '':
            email = get_email(website)
    except:
        pass

    print(f"URL: {url}")
    print(f"Name: {name}")
    print(f"Address: {address}")
    print(f"Website: {website}")
    print(f"Phone: {phone}")
    print(f"Email: {email}")

    write_data = [url, name, address, website, phone, email]
    xl_write(write_data)

driver = driver_define()
urls_filename = 'urls.txt'
urls = [line.strip('\n') for line in open(urls_filename).readlines()]

for url in urls:
    try:
        scrape_data(driver, url)
    except Exception as e:
        print(f"Error occurred while scraping {url}: {e}")
        print(traceback.format_exc())

driver.quit()